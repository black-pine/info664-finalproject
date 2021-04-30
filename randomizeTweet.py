import json
import random
from openpyxl import load_workbook
import datetime
import tweepy


wb = load_workbook('marvelData.xlsx')

# select random record from the current month
year = random.randint(2003, 2020)
month = datetime.datetime.now().month
try:
	sheet = wb[f"{year}_{month}"]
except:
	# force a valid year/month combination by removing incomplete years
	year = random.randint(2004, 2019)
	sheet = wb[f"{year}_{month}"]
	
row = random.randint(2, 5)
marvelId = sheet.cell(row = row, column = 14).value
# check for edge case of no API record found / saved
while not marvelId:
	row = random.randint(2, 5)
file = f"records/{marvelId}.json"

with open(file, 'r') as comic:
	data = json.load(comic)

	title = data['title']
	imgPath = data['images'][0]['path']
	imgExt = '.' + data['images'][0]['extension']
	for link in data['urls']:
		if link['type'] == 'detail':
			url = link['url']
	creators = {}
	for creator in data['creators']['items']:
		creators[creator['name']] = creator['role']

# build text for tweet
text = ''
yearDiff = datetime.datetime.now().year - year
if yearDiff == 1:
	text = "Last year on this month, "
else:
	text = f"{yearDiff} years ago on this month, "
text += title
if row == 2:
	text += ' was the top selling issue published by Marvel Comics.'
elif row == 3:
	text += ' was the top selling first issue published by Marvel Comics.'
elif row == 4:
	text += ' was the most reordered issue published by Marvel Comics.'
else:
	delta = '{:,}'.format(int(sheet.cell(row = row, column = 13).value))
	text += ' sold ' + delta + ' more copies than the previous issue.'

creatorText = title + ' Creative team: '
for creator in creators:
	creatorText += f"{creator} ({creators[creator]}), "
# remove ', ' from after the last appended creator
creatorText = creatorText[:len(creatorText)-2]

# twitter integration
apiKey = 'xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
secrcetAPI = 'xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
accessKey = 'xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
secretAccess = 'xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'

auth = tweepy.OAuthHandler(apiKey, secrcetAPI)
auth.set_access_token(accessKey, secretAccess)
api = tweepy.API(auth)

tweetText = text + '\n\n' + creatorText + '\n\n' + url
try:
	api.update_status(tweetText)
# remove creator text if tweet is too long
except:
	tweetText = text + '\n\n' + url
	api.update_status(tweetText)