import requests
import re
import json
from openpyxl import load_workbook
import hashlib
import time

titlePattern = re.compile(r'(.*)\s\(')

baseURL = 'https://gateway.marvel.com/'
reqURL = baseURL + '/v1/public/comics'
pubkey = 'ce0706691dc3583b90f39fd398cadf77'
pvtkey = '73be6d4a9e7fd520169cef87b17c1aa5e81b74e5'
ts = time.time()
hashval = f"{ts}{pvtkey}{pubkey}"
hashkey = hashlib.md5(hashval.encode()).hexdigest()

paramDiamond = {
	'apikey': pubkey,
	'hash': hashkey,
	'ts': ts,
	'noVariants': 'true',
	'diamondCode': ''
}
paramTitle = {
	'apikey': pubkey,
	'hash': hashkey,
	'ts': ts,
	'noVariants': 'true',
	'orderBy': 'onsaleDate',
	'title': '',
	'issueNumber': ''
}

pulledItems = set()
pulledComics = {}

wb = load_workbook('marvelData.xlsx')

for year in range(2003, 2021):
	for month in range(1, 13):
		# adjustments to get only range [3/2003, 3/2020]
		if year == 2003 and month < 3:
			continue
		if year == 2020 and month > 3:
			break

		sheet = wb[f"{year}_{month}"]
		sheet['N1'] = 'marvel id'
		for row in range(2, 6):
			diamondCode = sheet.cell(row = row, column = 6).value
			if diamondCode not in pulledItems:
				pulledItems.add(diamondCode)
				paramDiamond['diamondCode'] = diamondCode
				req = requests.get(reqURL, params=paramDiamond)
				data = json.loads(req.text)

				# try searching by title & issue number if diamond code is not listed in API
				if data['code'] == 200 and data['data']['total'] == 0:
					paramTitle['title'] = titlePattern.search(sheet.cell(row = row, column = 9).value).group(1)
					paramTitle['issueNumber'] = sheet.cell(row = row, column = 10).value
					req = requests.get(reqURL, params=paramTitle)
					data = json.loads(req.text)

				# save results as individual files
				if data['code'] == 200:
					if data['data']['total'] != 1:
						for comic in data['data']['results']:
							for date in comic['dates']:
								try:
									if date['type'] == 'onsaleDate' and date['date'].startswith(str(year)):
										filename = f"{comic['id']}.json"
										json.dump(comic, open('records/'+filename, 'w'), indent=4)
										sheet.cell(row = row, column = 14).value = comic['id']
								except:
									continue
					else:
						for comic in data['data']['results']:
							filename = f"{comic['id']}.json"
							json.dump(comic, open('records/'+filename, 'w'), indent=4)
							sheet.cell(row = row, column = 14).value = comic['id']

wb.save('marvelData.xlsx')